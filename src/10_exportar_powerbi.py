"""
10_exportar_powerbi.py
======================
Genera POWERBI_CABATECH_MASTER.xlsx con 3 hojas:
  - Master      : 168 subzonas x 87 cols, tabla Excel "tbl_subzonas"
  - Barrios     : 48 filas agregadas, tabla Excel "tbl_barrios"
  - Diccionario : metadata de columnas (nombre, tipo, descripcion, fuente)

Input : data/output/TABLEAU_CABATECH_MASTER.csv
Output: data/output/POWERBI_CABATECH_MASTER.xlsx
"""

import sys
import io
# Forzar UTF-8 en stdout para compatibilidad con terminales Windows (cp1252)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

# ─────────────────────────────────────────────
# 0. CONFIGURACIÓN
# ─────────────────────────────────────────────
INPUT_CSV   = os.path.join("data", "output", "TABLEAU_CABATECH_MASTER.csv")
OUTPUT_XLSX = os.path.join("data", "output", "POWERBI_CABATECH_MASTER.xlsx")

# ─────────────────────────────────────────────
# 1. CARGAR DATOS
# ─────────────────────────────────────────────
print("→ Cargando CSV master...")
df = pd.read_csv(INPUT_CSV, encoding="utf-8-sig")
print(f"✓ Dataset cargado: {df.shape[0]} subzonas × {df.shape[1]} columnas")
assert df.shape[0] == 168, f"⚠ Se esperaban 168 subzonas, se encontraron {df.shape[0]}"
assert df["barrio"].nunique() == 48, f"⚠ Se esperaban 48 barrios, se encontraron {df['barrio'].nunique()}"

# ─────────────────────────────────────────────
# 2. HOJA BARRIOS — agregado a nivel barrio
# ─────────────────────────────────────────────
print("→ Generando hoja Barrios (48 filas)...")

agg_dict = {
    # Precio
    "precio_m2":           "mean",
    "precio_venta_m2":     "mean",
    "precio_m2_zonaprop":  "mean",
    # Rentabilidad
    "roi":                 "mean",
    "cap_rate":            "mean",
    "sharpe":              "mean",
    "rentabilidad":        "mean",
    "payback":             "mean",
    # Score
    "score_inversion":     "mean",
    "score_transporte":    "mean",
    "score_fot":           "mean",
    "score_densidad":      "mean",
    "score_equipamiento":  "mean",
    "score_absorcion":     "mean",
    # Transporte (Capa 1)
    "dist_subte_m":        "mean",
    "cant_estaciones_800m":"mean",
    "subte_5min":          "sum",   # cantidad de subzonas a 5 min
    "subte_10min":         "sum",
    # FOT/FOS (Capa 2)
    "fot_promedio":        "mean",
    "fos_promedio":        "mean",
    "m2_edif_estimado":    "mean",
    # Densidad (Capa 3)
    "poblacion":           "sum",
    "densidad_pob_km2":    "mean",
    "pct_nbi":             "mean",
    "area_km2":            "sum",
    # POIs (Capa 4)
    "idx_equipamiento":    "mean",
    "poi_escuelas":        "sum",
    "poi_hospitales":      "sum",
    "poi_restaurantes":    "sum",
    "poi_comercios":       "sum",
    "poi_parques":         "sum",
    # Absorción (Capa 5)
    "stock_avisos":        "sum",
    # Subzonas por barrio
    "subzona":             "count",
}

df_barrios = df.groupby("barrio", as_index=False).agg(agg_dict)
df_barrios.rename(columns={"subzona": "n_subzonas"}, inplace=True)

# Redondear floats a 2 decimales
float_cols = df_barrios.select_dtypes(include="float64").columns
df_barrios[float_cols] = df_barrios[float_cols].round(2)

print(f"✓ Barrios generados: {df_barrios.shape[0]} filas × {df_barrios.shape[1]} columnas")

# ─────────────────────────────────────────────
# 3. DICCIONARIO DE COLUMNAS
# ─────────────────────────────────────────────
print("→ Construyendo diccionario de columnas...")

# (columna, tipo_friendly, descripción, fuente)
DICCIONARIO = [
    # ── GEOGRÁFICAS / BASE ──────────────────────────────────────────────────
    ("barrio",              "Texto",   "Nombre oficial del barrio CABA",                                    "Base"),
    ("subzona",             "Texto",   "Nombre de la subzona PropTech (3-4 por barrio)",                    "Base"),
    ("lat",                 "Decimal", "Latitud del centroide de la subzona",                               "Base"),
    ("lon",                 "Decimal", "Longitud del centroide de la subzona",                              "Base"),
    # ── CATEGÓRICAS ─────────────────────────────────────────────────────────
    ("categoria",           "Texto",   "Categoría de inversión derivada (A/B/C/D)",                        "Base"),
    ("categoria_precio",    "Texto",   "Premium / Medio / Económico según precio_m2",                      "Base"),
    ("categoria_score",     "Texto",   "Alto / Medio / Bajo según score_inversion",                        "Base"),
    ("perfil",              "Texto",   "Perfil urbano (residencial, mixto, comercial, etc.)",               "Base"),
    ("emergente",           "Entero",  "Flag 0/1: subzona con potencial de revalorización",                 "Base"),
    ("clase_gentrif",       "Texto",   "Clase de proceso de gentrificación",                                "Base"),
    ("zona_cur",            "Texto",   "Zona urbanística según CUR GCBA",                                   "Base"),
    # ── PRECIOS / MERCADO ───────────────────────────────────────────────────
    ("precio_m2",           "Entero",  "Precio de terreno en USD/m² (fuente PropTech)",                    "Base"),
    ("precio_venta_m2",     "Entero",  "Precio de venta estimado en USD/m² construido",                    "Base"),
    ("rentabilidad",        "Decimal", "Rentabilidad bruta estimada (%)",                                   "Base"),
    # ── MÉTRICAS FINANCIERAS ────────────────────────────────────────────────
    ("roi",                 "Decimal", "Retorno sobre inversión (%)",                                       "Base"),
    ("cap_rate",            "Decimal", "Cap Rate: NOI / precio_compra × 100",                               "Base"),
    ("sharpe",              "Decimal", "Sharpe Ratio: retorno ajustado por riesgo",                         "Base"),
    ("volatilidad",         "Decimal", "Volatilidad anualizada del precio/m²",                              "Base"),
    ("var95",               "Decimal", "Value at Risk al 95%: pérdida máxima estimada",                     "Base"),
    ("prob_positivo",       "Entero",  "Probabilidad de ROI positivo en simulación Monte Carlo (%)",        "Base"),
    ("payback",             "Decimal", "Años para recuperar la inversión",                                  "Base"),
    ("proyeccion",          "Decimal", "Valor proyectado del activo a 5 años (USD)",                        "Base"),
    ("margen",              "Decimal", "Margen neto del proyecto sobre costos (%)",                         "Base"),
    ("costo_total",         "Entero",  "Costo total estimado del proyecto (USD)",                           "Base"),
    ("ingresos",            "Entero",  "Ingresos totales estimados por venta (USD)",                        "Base"),
    # ── NORMATIVA / EDIFICABILIDAD ───────────────────────────────────────────
    ("pisos",               "Texto",   "Rango de pisos permitidos según normativa",                         "Base"),
    ("n_pisos",             "Entero",  "Número de pisos máximo estimado",                                   "Base"),
    ("altura_max",          "Decimal", "Altura máxima estimada (m) según normativa base",                   "Base"),
    ("m2_vendibles",        "Entero",  "M² vendibles estimados en el proyecto tipo",                        "Base"),
    ("ff5",                 "Decimal", "Factor de flujo futuro a 5 años",                                   "Base"),
    # ── SCORES ──────────────────────────────────────────────────────────────
    ("score_inversion",     "Decimal", "Score de inversión compuesto 0-100 (6 componentes)",                "Base"),
    ("score_original",      "Entero",  "Score base del dataset PropTech original (0-100)",                  "Base"),
    ("score_delta",         "Decimal", "Diferencia score_inversion vs score_original",                      "Base"),
    ("score_transporte",    "Decimal", "Componente transporte del score (0-100)",                           "Capa 1"),
    ("score_fot",           "Decimal", "Componente edificabilidad FOT del score (0-100)",                   "Capa 2"),
    ("score_densidad",      "Decimal", "Componente densidad poblacional del score (0-100)",                 "Capa 3"),
    ("score_equipamiento",  "Decimal", "Componente equipamiento urbano del score (0-100)",                  "Capa 4"),
    ("score_absorcion",     "Decimal", "Componente liquidez de mercado del score (0-100)",                  "Capa 5"),
    ("score",               "Entero",  "Score entero redondeado (alias de score_inversion × 100)",          "Base"),
    ("score_desarrollo",    "Entero",  "Score de potencial de desarrollo urbano",                           "Base"),
    # ── ÍNDICES DERIVADOS ────────────────────────────────────────────────────
    ("accesibilidad",       "Entero",  "Índice de accesibilidad urbana (0-100)",                            "Base"),
    ("idx_gentrificacion",  "Entero",  "Índice compuesto de gentrificación (0-100)",                        "Base"),
    ("gv",                  "Entero",  "Componente gentrificación: valor inmobiliario",                     "Base"),
    ("gp",                  "Entero",  "Componente gentrificación: perfil socioeconómico",                  "Base"),
    ("gc",                  "Entero",  "Componente gentrificación: conectividad",                           "Base"),
    ("gt",                  "Entero",  "Componente gentrificación: transformación urbana",                  "Base"),
    # ── CAPA 1: TRANSPORTE ───────────────────────────────────────────────────
    ("dist_subte_m",        "Decimal", "Distancia en metros a la estación de subte más cercana",            "Capa 1"),
    ("cant_estaciones_800m","Entero",  "Cantidad de estaciones de subte en radio de 800 m",                 "Capa 1"),
    ("linea_subte_cercana", "Texto",   "Línea de subte más cercana (A/B/C/D/E/H)",                          "Capa 1"),
    ("estacion_cercana",    "Texto",   "Nombre de la estación de subte más cercana",                        "Capa 1"),
    ("subte_5min",          "Entero",  "Flag 0/1: hay estación de subte en radio de 5 min a pie",           "Capa 1"),
    ("subte_10min",         "Entero",  "Flag 0/1: hay estación de subte en radio de 10 min a pie",          "Capa 1"),
    # ── CAPA 2: FOT / FOS ────────────────────────────────────────────────────
    ("fot_promedio",        "Decimal", "Factor de Ocupación Total promedio de parcelas en la subzona",      "Capa 2"),
    ("fot_mediana",         "Decimal", "FOT mediana de la subzona",                                         "Capa 2"),
    ("fot_maximo",          "Decimal", "FOT máximo registrado en la subzona",                               "Capa 2"),
    ("fos_promedio",        "Decimal", "Factor de Ocupación del Suelo promedio",                            "Capa 2"),
    ("fos_maximo",          "Decimal", "FOS máximo registrado en la subzona",                               "Capa 2"),
    ("altura_max_promedio", "Decimal", "Altura máxima promedio según normativa CUR (m)",                    "Capa 2"),
    ("altura_max_absoluta", "Decimal", "Altura máxima absoluta registrada en la subzona (m)",               "Capa 2"),
    ("dist_cur_dominante",  "Texto",   "Distrito CUR predominante en la subzona",                           "Capa 2"),
    ("n_distritos_cur",     "Decimal", "Cantidad de distritos CUR presentes en la subzona",                 "Capa 2"),
    ("m2_edif_estimado",    "Decimal", "M² edificables estimados según FOT promedio × área",                "Capa 2"),
    # ── CAPA 3: DENSIDAD / CENSO ─────────────────────────────────────────────
    ("poblacion",           "Decimal", "Población total de la subzona (INDEC Censo 2022)",                  "Capa 3"),
    ("viviendas",           "Decimal", "Cantidad de viviendas particulares (Censo 2022)",                   "Capa 3"),
    ("hogares",             "Decimal", "Cantidad de hogares (Censo 2022)",                                  "Capa 3"),
    ("hogares_nbi",         "Decimal", "Hogares con al menos una NBI (Censo 2022)",                         "Capa 3"),
    ("densidad_pob_km2",    "Decimal", "Densidad poblacional en hab/km²",                                   "Capa 3"),
    ("densidad_viv_km2",    "Decimal", "Densidad de viviendas en viv/km²",                                  "Capa 3"),
    ("pct_nbi",             "Decimal", "Porcentaje de hogares con NBI (%)",                                 "Capa 3"),
    ("pob_por_hogar",       "Decimal", "Promedio de personas por hogar",                                    "Capa 3"),
    ("area_km2",            "Decimal", "Área de la subzona en km²",                                         "Capa 3"),
    # ── CAPA 4: POIs / EQUIPAMIENTO ─────────────────────────────────────────
    ("idx_equipamiento",    "Decimal", "Índice de equipamiento urbano (0-100) basado en POIs",              "Capa 4"),
    ("poi_escuelas",        "Entero",  "Cantidad de escuelas en radio de 1 km (OSM)",                       "Capa 4"),
    ("poi_universidades",   "Entero",  "Cantidad de universidades en radio de 1 km (OSM)",                  "Capa 4"),
    ("poi_hospitales",      "Entero",  "Cantidad de hospitales en radio de 1 km (OSM)",                     "Capa 4"),
    ("poi_clinicas",        "Entero",  "Cantidad de clínicas en radio de 1 km (OSM)",                       "Capa 4"),
    ("poi_restaurantes",    "Entero",  "Cantidad de restaurantes en radio de 1 km (OSM)",                   "Capa 4"),
    ("poi_cafes",           "Entero",  "Cantidad de cafeterías en radio de 1 km (OSM)",                     "Capa 4"),
    ("poi_comercios",       "Entero",  "Cantidad de comercios en radio de 1 km (OSM)",                      "Capa 4"),
    ("poi_bancos",          "Entero",  "Cantidad de bancos/cajeros en radio de 1 km (OSM)",                 "Capa 4"),
    ("poi_farmacias",       "Entero",  "Cantidad de farmacias en radio de 1 km (OSM)",                      "Capa 4"),
    ("poi_parques",         "Entero",  "Cantidad de plazas y parques en radio de 1 km (OSM)",               "Capa 4"),
    ("poi_gastronomia",     "Entero",  "Total de POIs gastronómicos (restaurantes + cafés)",                 "Capa 4"),
    # ── CAPA 5: ABSORCIÓN DE MERCADO ────────────────────────────────────────
    ("stock_avisos",        "Decimal", "Cantidad de avisos activos en ZonaProp (snapshot)",                 "Capa 5"),
    ("precio_m2_zonaprop",  "Decimal", "Precio/m² promedio según avisos ZonaProp (USD/m²)",                 "Capa 5"),
    ("precio_usd_zonaprop", "Decimal", "Precio total promedio de aviso en ZonaProp (USD)",                  "Capa 5"),
    ("sup_prom_zonaprop",   "Decimal", "Superficie promedio de los inmuebles en avisos ZonaProp (m²)",      "Capa 5"),
]

df_dic = pd.DataFrame(DICCIONARIO, columns=["columna", "tipo", "descripcion", "fuente"])

# Verificar cobertura
cols_en_dic = set(df_dic["columna"])
cols_en_df  = set(df.columns)
faltantes   = cols_en_df - cols_en_dic
sobrantes   = cols_en_dic - cols_en_df
if faltantes:
    print(f"  ⚠ Columnas del CSV sin entrada en diccionario: {sorted(faltantes)}")
if sobrantes:
    print(f"  ⚠ Entradas en diccionario sin columna en CSV:  {sorted(sobrantes)}")
print(f"✓ Diccionario: {len(df_dic)} entradas ({len(cols_en_dic & cols_en_df)} cols mapeadas)")

# ─────────────────────────────────────────────
# 4. ESCRIBIR XLSX
# ─────────────────────────────────────────────
print(f"→ Escribiendo {OUTPUT_XLSX}...")

with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
    df.to_excel(writer,         sheet_name="Master",      index=False)
    df_barrios.to_excel(writer, sheet_name="Barrios",     index=False)
    df_dic.to_excel(writer,     sheet_name="Diccionario", index=False)

print("✓ XLSX escrito (sin tablas aún)")

# ─────────────────────────────────────────────
# 5. APLICAR FORMATO DE TABLAS EXCEL (openpyxl)
# ─────────────────────────────────────────────
print("→ Aplicando formato de tablas Excel...")

wb = load_workbook(OUTPUT_XLSX)

# ── Función helper ──────────────────────────────────────────────────────────
def aplicar_tabla(ws, nombre_tabla, estilo="TableStyleMedium9"):
    """Convierte el rango de datos de una hoja en tabla Excel con nombre."""
    max_col_letra = get_column_letter(ws.max_column)
    max_row       = ws.max_row
    ref           = f"A1:{max_col_letra}{max_row}"

    tabla = Table(displayName=nombre_tabla, ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name=estilo,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabla)

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    print(f"  ✓ Tabla '{nombre_tabla}' creada en hoja '{ws.title}' — rango {ref}")

# ── Hoja Master ─────────────────────────────────────────────────────────────
ws_master = wb["Master"]
aplicar_tabla(ws_master, "tbl_subzonas", estilo="TableStyleMedium2")

# Freeze pane en la fila 1 (encabezados fijos)
ws_master.freeze_panes = "A2"

# ── Hoja Barrios ────────────────────────────────────────────────────────────
ws_barrios = wb["Barrios"]
aplicar_tabla(ws_barrios, "tbl_barrios", estilo="TableStyleMedium7")
ws_barrios.freeze_panes = "A2"

# ── Hoja Diccionario — tabla sin estilo bandeado (lectura) ──────────────────
ws_dic = wb["Diccionario"]
aplicar_tabla(ws_dic, "tbl_diccionario", estilo="TableStyleLight1")

# ── Guardar ─────────────────────────────────────────────────────────────────
wb.save(OUTPUT_XLSX)
print(f"✓ Tablas aplicadas y archivo guardado")

# ─────────────────────────────────────────────
# 6. VALIDACIÓN FINAL
# ─────────────────────────────────────────────
print("\n── VALIDACIÓN ─────────────────────────────────────────────")
wb_v = load_workbook(OUTPUT_XLSX)

ws_m = wb_v["Master"]
ws_b = wb_v["Barrios"]
ws_d = wb_v["Diccionario"]

print(f"  Master      : {ws_m.max_row - 1} filas × {ws_m.max_column} cols  (esperado 168 × 87)")
print(f"  Barrios     : {ws_b.max_row - 1} filas × {ws_b.max_column} cols  (esperado  48 × {df_barrios.shape[1]})")
print(f"  Diccionario : {ws_d.max_row - 1} entradas")

tablas_master   = [t.displayName for t in ws_m.tables.values()]
tablas_barrios  = [t.displayName for t in ws_b.tables.values()]
tablas_dic      = [t.displayName for t in ws_d.tables.values()]
print(f"  Tablas Excel detectadas:")
print(f"    Master      → {tablas_master}")
print(f"    Barrios     → {tablas_barrios}")
print(f"    Diccionario → {tablas_dic}")

assert "tbl_subzonas"   in tablas_master,  "❌ tbl_subzonas no encontrada"
assert "tbl_barrios"    in tablas_barrios, "❌ tbl_barrios no encontrada"
assert "tbl_diccionario" in tablas_dic,    "❌ tbl_diccionario no encontrada"

file_kb = os.path.getsize(OUTPUT_XLSX) / 1024
print(f"\n✓ {OUTPUT_XLSX} generado correctamente ({file_kb:.1f} KB)")
print("  Power BI puede conectar directamente a las 3 tablas:")
print("  → tbl_subzonas   (168 subzonas, granularidad máxima)")
print("  → tbl_barrios    (48 barrios, agregado)")
print("  → tbl_diccionario (metadata de columnas)")
